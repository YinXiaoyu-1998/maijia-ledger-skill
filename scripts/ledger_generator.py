#!/usr/bin/env python3
"""Generate a food purchase ledger from receipt workbooks."""

from __future__ import annotations

import argparse
import sys
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
    module="openpyxl.styles.stylesheet",
)

LEDGER_HEADERS = [
    "序号",
    "收货单号",
    "门店名",
    "食品名称",
    "规格",
    "进货数量",
    "进货金额",
    "生产日期或生产批号",
    "保质期",
    "供货单位名称",
    "供货单位地址",
    "供货单位联系方式",
    "进货日期",
]

REQUIRED_DETAIL_HEADERS = ["序号", "物品名称", "规格型号", "收货数量", "收货金额"]
REQUIRED_META_HEADERS = ["单据号", "订货机构", "送货机构", "收货日期"]
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
SUPPLIER_CATALOG_FILENAMES = ("supplier_catalog.xlsx", "供应商档案.xlsx")


@dataclass(frozen=True)
class ReceiptItem:
    """One item row parsed from a receipt workbook."""

    document_no: str
    store_name: str
    received_date: date | str | None
    supplier_name: str
    item_name: str
    specification: str
    quantity: Any
    amount: Any


@dataclass
class RunSummary:
    """Counters and warnings for one generation run."""

    input_files: int = 0
    parsed_files: int = 0
    skipped_duplicate_files: int = 0
    skipped_invalid_files: int = 0
    appended_rows: int = 0
    supplier_catalog_used: bool = False
    created_ledgers: int = 0
    updated_ledgers: int = 0
    output_files: list[Path] | None = None
    warnings: list[str] | None = None

    def add_warning(self, message: str) -> None:
        if self.warnings is None:
            self.warnings = []
        if message not in self.warnings:
            self.warnings.append(message)

    def add_output_file(self, path: Path) -> None:
        if self.output_files is None:
            self.output_files = []
        if path not in self.output_files:
            self.output_files.append(path)


def parse_args(argv: list[str]) -> argparse.Namespace:
    """Parse command-line arguments."""

    parser = argparse.ArgumentParser(
        description="把配送收货单 Excel 转换为食品经营单位进货台账。"
    )
    parser.add_argument("-dir", dest="input_dir", type=Path, help="收货单据文件夹")
    parser.add_argument("-new_file", dest="new_file", type=Path, help="单个收货单据文件")
    parser.add_argument(
        "-original_file",
        "-origininal_file",
        dest="original_file",
        type=Path,
        help="已有台账文件；提供后默认原地追加更新",
    )
    parser.add_argument(
        "-supplier_catalog",
        dest="supplier_catalog",
        type=Path,
        help="可选供应商档案文件；未提供时自动寻找 supplier_catalog.xlsx 或 供应商档案.xlsx",
    )
    parser.add_argument(
        "-output",
        dest="output",
        type=Path,
        help="可选输出路径；默认按门店自动创建或更新 台账_<门店名>.xlsx；多门店时可提供输出文件夹",
    )

    args = parser.parse_args(argv)
    if args.input_dir is None and args.new_file is None:
        parser.error("请至少提供 -dir 或 -new_file。")
    return args


def is_receipt_workbook(path: Path) -> bool:
    """Return whether a path looks like a real receipt workbook."""

    if not path.is_file():
        return False
    if path.suffix.lower() not in EXCEL_SUFFIXES:
        return False
    return not (path.name.startswith("~$") or path.name.startswith(".~"))


def collect_input_files(input_dir: Path | None, new_file: Path | None) -> list[Path]:
    """Collect receipt workbook paths from a file and/or directory."""

    files: list[Path] = []
    if input_dir is not None:
        if not input_dir.is_dir():
            raise FileNotFoundError(f"找不到文件夹: {input_dir}")
        files.extend(path for path in sorted(input_dir.iterdir()) if is_receipt_workbook(path))

    if new_file is not None:
        if not new_file.is_file():
            raise FileNotFoundError(f"找不到文件: {new_file}")
        if is_receipt_workbook(new_file):
            files.append(new_file)

    seen: set[Path] = set()
    unique_files: list[Path] = []
    for file_path in files:
        resolved = file_path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique_files.append(file_path)
    return unique_files


def normalized_text(value: Any) -> str:
    """Convert a cell value to a stripped display string."""

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_date(value: Any) -> date | str | None:
    """Parse common date cell values while preserving unknown values."""

    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = normalized_text(value)
    if not text:
        return None

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return text


def find_label_value(ws: Worksheet, label: str, max_rows: int = 8) -> Any:
    """Find a label in the top part of a sheet and return the next non-empty cell."""

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows)):
        for index, cell in enumerate(row):
            if normalized_text(cell.value) != label:
                continue
            for next_cell in row[index + 1 :]:
                if normalized_text(next_cell.value):
                    return next_cell.value
            return None
    return None


def find_detail_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """Find the receipt detail header row and return a header-to-column map."""

    for row_number in range(1, min(ws.max_row, 25) + 1):
        headers: dict[str, int] = {}
        for column_number in range(1, ws.max_column + 1):
            value = normalized_text(ws.cell(row_number, column_number).value)
            if value:
                headers[value] = column_number
        if all(header in headers for header in REQUIRED_DETAIL_HEADERS):
            return row_number, headers
    raise ValueError("找不到物品明细表头。")


def read_supplier_phones(catalog_path: Path) -> dict[str, str]:
    """Read supplier phone numbers from the supplier catalog workbook."""

    wb = load_workbook(catalog_path, data_only=True, read_only=False)
    ws = wb.active
    header_row = find_header_row(ws, ["供应商名称", "联系人电话"], max_rows=20)
    header_map = row_header_map(ws, header_row)
    phone_by_name: dict[str, str] = {}

    name_columns = [
        header_map[column_name]
        for column_name in ("供应商名称", "供应商简称")
        if column_name in header_map
    ]
    phone_column = header_map["联系人电话"]

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        phone = normalized_text(row[phone_column - 1] if len(row) >= phone_column else None)
        for name_column in name_columns:
            supplier_name = normalized_text(row[name_column - 1] if len(row) >= name_column else None)
            if supplier_name and supplier_name not in phone_by_name:
                phone_by_name[supplier_name] = phone

    return phone_by_name


def find_header_row(ws: Worksheet, required_headers: list[str], max_rows: int = 25) -> int:
    """Find a row containing all required headers."""

    for row_number in range(1, min(ws.max_row, max_rows) + 1):
        row_values = {
            normalized_text(ws.cell(row_number, column_number).value)
            for column_number in range(1, ws.max_column + 1)
        }
        if all(header in row_values for header in required_headers):
            return row_number
    raise ValueError(f"找不到表头: {', '.join(required_headers)}")


def row_header_map(ws: Worksheet, row_number: int) -> dict[str, int]:
    """Map non-empty header labels in a row to one-based column numbers."""

    return {
        normalized_text(ws.cell(row_number, column_number).value): column_number
        for column_number in range(1, ws.max_column + 1)
        if normalized_text(ws.cell(row_number, column_number).value)
    }


def parse_receipt(path: Path) -> tuple[str, list[ReceiptItem]]:
    """Parse one receipt workbook into ledger rows."""

    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb.active

    meta = {header: find_label_value(ws, header) for header in REQUIRED_META_HEADERS}
    missing_meta = [header for header, value in meta.items() if normalized_text(value) == ""]
    if missing_meta:
        raise ValueError(f"缺少单据头字段: {', '.join(missing_meta)}")

    header_row, headers = find_detail_header(ws)
    document_no = normalized_text(meta["单据号"])
    store_name = normalized_text(meta["订货机构"])
    supplier_name = normalized_text(meta["送货机构"])
    received_date = parse_date(meta["收货日期"])

    items: list[ReceiptItem] = []
    sequence_column = headers["序号"]
    for row_number in range(header_row + 1, ws.max_row + 1):
        sequence_value = normalized_text(ws.cell(row_number, sequence_column).value)
        if sequence_value == "合计":
            break
        item_name = normalized_text(ws.cell(row_number, headers["物品名称"]).value)
        if not item_name:
            continue
        items.append(
            ReceiptItem(
                document_no=document_no,
                store_name=store_name,
                received_date=received_date,
                supplier_name=supplier_name,
                item_name=item_name,
                specification=normalized_text(ws.cell(row_number, headers["规格型号"]).value),
                quantity=ws.cell(row_number, headers["收货数量"]).value,
                amount=ws.cell(row_number, headers["收货金额"]).value,
            )
        )

    return document_no, items


def new_ledger_workbook() -> tuple[Workbook, Worksheet, int, dict[str, int]]:
    """Create a new ledger workbook with standard headers."""

    wb = Workbook()
    ws = wb.active
    ws.title = "台账"
    for column_number, header in enumerate(LEDGER_HEADERS, start=1):
        ws.cell(1, column_number, header)
    style_header(ws, 1)
    return wb, ws, 1, row_header_map(ws, 1)


def prepare_ledger(
    ledger_file: Path | None, create_if_missing: bool = False
) -> tuple[Workbook, Worksheet, int, dict[str, int]]:
    """Open an existing ledger or create a new one with standard headers."""

    if ledger_file is None:
        return new_ledger_workbook()

    if ledger_file.is_file():
        wb = load_workbook(ledger_file)
        ws = wb.active
        header_row = find_header_row(ws, ["收货单号"], max_rows=20)
        header_map = row_header_map(ws, header_row)
        ensure_ledger_headers(ws, header_row, header_map)
        return wb, ws, header_row, row_header_map(ws, header_row)

    if create_if_missing:
        return new_ledger_workbook()

    raise FileNotFoundError(f"找不到已有台账: {ledger_file}")


def ensure_ledger_headers(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> None:
    """Append missing expected ledger headers to an existing ledger."""

    next_column = ws.max_column + 1
    for header in LEDGER_HEADERS:
        if header in header_map:
            continue
        ws.cell(header_row, next_column, header)
        header_map[header] = next_column
        next_column += 1
    style_header(ws, header_row)


def style_header(ws: Worksheet, header_row: int) -> None:
    """Apply a simple, readable style to the ledger header row."""

    fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[header_row]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def existing_document_numbers(
    ws: Worksheet, header_row: int, header_map: dict[str, int]
) -> set[str]:
    """Read receipt document numbers already present in a ledger."""

    document_column = header_map.get("收货单号")
    if document_column is None:
        return set()
    document_numbers: set[str] = set()
    for row_number in range(header_row + 1, ws.max_row + 1):
        document_no = normalized_text(ws.cell(row_number, document_column).value)
        if document_no:
            document_numbers.add(document_no)
    return document_numbers


def next_sequence_number(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> int:
    """Return the next ledger sequence number."""

    sequence_column = header_map.get("序号")
    if sequence_column is None:
        return 1
    max_sequence = 0
    for row_number in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row_number, sequence_column).value
        if isinstance(value, int):
            max_sequence = max(max_sequence, value)
        elif isinstance(value, float) and value.is_integer():
            max_sequence = max(max_sequence, int(value))
        elif isinstance(value, str) and value.strip().isdigit():
            max_sequence = max(max_sequence, int(value.strip()))
    return max_sequence + 1


def append_items(
    ws: Worksheet,
    header_map: dict[str, int],
    items: Iterable[ReceiptItem],
    phone_by_supplier: dict[str, str],
    start_sequence: int,
) -> int:
    """Append receipt items to the ledger sheet and return the next sequence number."""

    sequence = start_sequence
    for item in items:
        row_number = ws.max_row + 1
        values = {
            "序号": sequence,
            "收货单号": item.document_no,
            "门店名": item.store_name,
            "食品名称": item.item_name,
            "规格": item.specification,
            "进货数量": item.quantity,
            "进货金额": item.amount,
            "生产日期或生产批号": "",
            "保质期": "",
            "供货单位名称": item.supplier_name,
            "供货单位地址": "",
            "供货单位联系方式": phone_by_supplier.get(item.supplier_name, ""),
            "进货日期": item.received_date,
        }
        for header, value in values.items():
            cell = ws.cell(row_number, header_map[header], value)
            if header == "进货日期" and isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
        sequence += 1
    return sequence


def format_ledger(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> None:
    """Apply final formatting to a ledger sheet."""

    widths = {
        "序号": 8,
        "收货单号": 18,
        "门店名": 24,
        "食品名称": 28,
        "规格": 20,
        "进货数量": 12,
        "进货金额": 12,
        "生产日期或生产批号": 22,
        "保质期": 12,
        "供货单位名称": 34,
        "供货单位地址": 22,
        "供货单位联系方式": 18,
        "进货日期": 14,
    }
    for header, width in widths.items():
        column_number = header_map.get(header)
        if column_number is not None:
            ws.column_dimensions[get_column_letter(column_number)].width = width

    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def output_path_for(original_file: Path | None, output: Path | None) -> Path:
    """Choose the destination ledger path."""

    if output is not None:
        return output
    if original_file is not None:
        return original_file
    return Path("台账.xlsx")


def safe_filename_part(value: str) -> str:
    """Return a filesystem-safe filename component."""

    cleaned = "".join("_" if char in '<>:"/\\|?*' or ord(char) < 32 else char for char in value)
    cleaned = cleaned.strip().strip(".")
    return cleaned or "未命名门店"


def default_store_ledger_filename(store_name: str) -> str:
    """Return the default ledger filename for one store."""

    return f"台账_{safe_filename_part(store_name)}.xlsx"


def output_path_for_store(store_name: str, output: Path | None, store_count: int) -> Path:
    """Choose the destination ledger path for one store in split-ledger mode."""

    default_name = default_store_ledger_filename(store_name)
    if output is None:
        return Path(default_name)
    if output.suffix.lower() in EXCEL_SUFFIXES:
        if store_count == 1:
            return output
        raise ValueError("多门店输入不能把 -output 指向单个 Excel 文件；请省略 -output 或提供输出文件夹。")
    return output / default_name


def candidate_catalog_dirs(input_dir: Path | None, new_file: Path | None) -> list[Path]:
    """Return likely directories that may contain a supplier catalog."""

    candidates = [Path.cwd()]
    if input_dir is not None:
        candidates.extend([input_dir, input_dir.parent])
    if new_file is not None:
        candidates.extend([new_file.parent, new_file.parent.parent])

    unique_candidates: list[Path] = []
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved not in seen:
            seen.add(resolved)
            unique_candidates.append(candidate)
    return unique_candidates


def resolve_supplier_catalog_path(
    catalog_path: Path | None, input_dir: Path | None, new_file: Path | None
) -> Path | None:
    """Return the supplier catalog to use, if one is available."""

    if catalog_path is not None:
        return catalog_path

    for candidate_dir in candidate_catalog_dirs(input_dir, new_file):
        for filename in SUPPLIER_CATALOG_FILENAMES:
            candidate = candidate_dir / filename
            if candidate.is_file():
                return candidate
    return None


def load_supplier_phones(args: argparse.Namespace, summary: RunSummary) -> dict[str, str]:
    """Load optional supplier phone lookup data."""

    supplier_catalog_path = resolve_supplier_catalog_path(
        args.supplier_catalog, args.input_dir, args.new_file
    )
    if supplier_catalog_path is None:
        summary.add_warning("未提供供应商目录，供货单位联系方式将留空。")
        return {}
    if supplier_catalog_path.is_file():
        summary.supplier_catalog_used = True
        return read_supplier_phones(supplier_catalog_path)

    summary.add_warning(f"供应商目录不存在，供货单位联系方式将留空: {supplier_catalog_path}")
    return {}


def append_receipts_to_ledger(
    source_file: Path,
    destination_file: Path,
    receipts: list[tuple[str, list[ReceiptItem]]],
    phone_by_supplier: dict[str, str],
    summary: RunSummary,
    create_if_missing: bool,
) -> int:
    """Append parsed receipts into one ledger file and return appended row count."""

    source_existed = source_file.is_file()
    destination_existed = destination_file.is_file()
    wb, ws, header_row, header_map = prepare_ledger(source_file, create_if_missing=create_if_missing)
    seen_document_numbers = existing_document_numbers(ws, header_row, header_map)
    sequence = next_sequence_number(ws, header_row, header_map)
    appended_rows = 0

    for document_no, items in receipts:
        if document_no in seen_document_numbers:
            summary.skipped_duplicate_files += 1
            continue

        summary.parsed_files += 1
        before_sequence = sequence
        sequence = append_items(ws, header_map, items, phone_by_supplier, sequence)
        added_rows = sequence - before_sequence
        appended_rows += added_rows
        summary.appended_rows += added_rows
        seen_document_numbers.add(document_no)

        if summary.supplier_catalog_used and items and items[0].supplier_name not in phone_by_supplier:
            summary.add_warning(f"供应商未匹配，联系方式留空: {items[0].supplier_name}")

    if appended_rows == 0 and source_existed and source_file == destination_file:
        return 0

    style_header(ws, header_row)
    format_ledger(ws, header_row, header_map)
    destination_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination_file)
    summary.add_output_file(destination_file)
    if source_existed or destination_existed:
        summary.updated_ledgers += 1
    else:
        summary.created_ledgers += 1
    return appended_rows


def parse_receipt_files(
    input_files: list[Path], summary: RunSummary
) -> list[tuple[str, list[ReceiptItem]]]:
    """Parse input files while collecting invalid-file warnings."""

    parsed_receipts: list[tuple[str, list[ReceiptItem]]] = []
    for input_file in input_files:
        try:
            document_no, items = parse_receipt(input_file)
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            summary.skipped_invalid_files += 1
            summary.add_warning(f"跳过 {input_file.name}: {exc}")
            continue
        if not items:
            summary.skipped_invalid_files += 1
            summary.add_warning(f"跳过 {input_file.name}: 没有可写入台账的物品明细。")
            continue
        parsed_receipts.append((document_no, items))
    return parsed_receipts


def generate_single_ledger(
    args: argparse.Namespace,
    receipts: list[tuple[str, list[ReceiptItem]]],
    phone_by_supplier: dict[str, str],
    summary: RunSummary,
) -> None:
    """Update one explicitly selected ledger file."""

    destination = output_path_for(args.original_file, args.output)
    source = args.original_file if args.original_file is not None else destination
    append_receipts_to_ledger(
        source,
        destination,
        receipts,
        phone_by_supplier,
        summary,
        create_if_missing=args.original_file is None,
    )


def generate_store_ledgers(
    args: argparse.Namespace,
    receipts: list[tuple[str, list[ReceiptItem]]],
    phone_by_supplier: dict[str, str],
    summary: RunSummary,
) -> None:
    """Create or update one ledger file per store."""

    receipts_by_store: dict[str, list[tuple[str, list[ReceiptItem]]]] = {}
    for document_no, items in receipts:
        store_name = items[0].store_name
        receipts_by_store.setdefault(store_name, []).append((document_no, items))

    store_count = len(receipts_by_store)
    for store_name, store_receipts in sorted(receipts_by_store.items()):
        destination = output_path_for_store(store_name, args.output, store_count)
        append_receipts_to_ledger(
            destination,
            destination,
            store_receipts,
            phone_by_supplier,
            summary,
            create_if_missing=True,
        )


def generate_ledger(args: argparse.Namespace) -> RunSummary:
    """Generate or update ledgers according to parsed arguments."""

    summary = RunSummary()
    input_files = collect_input_files(args.input_dir, args.new_file)
    summary.input_files = len(input_files)
    if not input_files:
        raise ValueError("没有找到可处理的 Excel 收货单据。")

    phone_by_supplier = load_supplier_phones(args, summary)
    receipts = parse_receipt_files(input_files, summary)
    if not receipts:
        raise ValueError("没有成功解析任何收货单据。")

    if args.original_file is not None:
        generate_single_ledger(args, receipts, phone_by_supplier, summary)
    else:
        generate_store_ledgers(args, receipts, phone_by_supplier, summary)

    return summary


def print_summary(summary: RunSummary) -> None:
    """Print a compact run summary."""

    print(f"输入文件: {summary.input_files}")
    print(f"成功处理: {summary.parsed_files}")
    print(f"追加行数: {summary.appended_rows}")
    print(f"跳过重复单据: {summary.skipped_duplicate_files}")
    print(f"跳过无效文件: {summary.skipped_invalid_files}")
    print(f"使用供应商目录: {'是' if summary.supplier_catalog_used else '否'}")
    print(f"新建台账: {summary.created_ledgers}")
    print(f"更新台账: {summary.updated_ledgers}")
    if summary.output_files:
        print("输出文件:")
        for output_file in summary.output_files:
            print(f"- {output_file}")
    if summary.warnings:
        print("提示:")
        for warning in summary.warnings:
            print(f"- {warning}")


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""

    args = parse_args(sys.argv[1:] if argv is None else argv)
    try:
        summary = generate_ledger(args)
    except Exception as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1
    print_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
